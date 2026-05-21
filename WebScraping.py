import requests
from bs4 import BeautifulSoup
import openpyxl
import time

BASE_URL = "https://books.toscrape.com/catalogue/"

RATING_MAP = {
    "One": 1,
    "Two": 2,
    "Three": 3,
    "Four": 4,
    "Five": 5
}

def scrape_page(url):
    response = requests.get(url)
    response.encoding = "utf-8"
    soup = BeautifulSoup(response.text, "html.parser")

    books = []
    articles = soup.select("article.product_pod")

    for article in articles:
        title = article.h3.a["title"]

        price = article.select_one("p.price_color").text.strip()
        price = price.encode("ascii", "ignore").decode().replace("£", "").strip()

        rating_word = article.p["class"][1]
        rating = RATING_MAP.get(rating_word, 0)

        availability = article.select_one("p.availability").text.strip()

        relative_url = article.h3.a["href"].replace("../", "")
        book_url = BASE_URL + relative_url

        books.append({
            "title": title,
            "price_gbp": price,
            "rating": rating,
            "availability": availability,
            "url": book_url
        })

    return books, soup


def get_next_page(soup):
    next_btn = soup.select_one("li.next a")
    if next_btn:
        href = next_btn["href"]
        # Fix: page-2.html direct use pannanum
        if "/" not in href:
            return BASE_URL + href
        return BASE_URL + href
    return None


def scrape_all_books(max_pages=None):
    all_books = []
    url = BASE_URL + "page-1.html"
    page_num = 1

    while url:
        print(f"Page {page_num} scraping... ({url})")
        books, soup = scrape_page(url)
        all_books.extend(books)
        print(f"  -> {len(books)} books found (Total: {len(all_books)})")

        if max_pages and page_num >= max_pages:
            break

        url = get_next_page(soup)
        page_num += 1
        time.sleep(0.3)

    return all_books


def save_to_excel(books, filename="books_data.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Books"

    # Header styling
    headers = ["Title", "Price (£)", "Rating (1-5)", "Availability", "URL"]
    ws.append(headers)

    # Bold header
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for i, book in enumerate(books):
        ws.append([
            book["title"],
            float(book["price_gbp"]) if book["price_gbp"] else 0,
            book["rating"],
            book["availability"],
            book["url"]
        ])
        # Alternate row color
        if i % 2 == 0:
            fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
            for col in range(1, 6):
                ws.cell(row=i+2, column=col).fill = fill

    # Column widths
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 60

    wb.save(filename)
    print(f"\n✅ {len(books)} books saved to '{filename}'")


if __name__ == "__main__":
    print("=" * 50)
    print("  Books to Scrape - Web Scraper")
    print("=" * 50)

    # 1000 books all pages - max_pages=None
    books = scrape_all_books(max_pages=None)

    save_to_excel(books, "books_data.xlsx")

    print("\n📊 Summary:")
    print(f"   Total Books     : {len(books)}")
    prices = [float(b['price_gbp']) for b in books if b['price_gbp']]
    if prices:
        print(f"   Avg Price       : £{sum(prices)/len(prices):.2f}")
        print(f"   Cheapest        : £{min(prices):.2f}")
        print(f"   Most Expensive  : £{max(prices):.2f}")